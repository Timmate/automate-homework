#! python3
#
# NAME         : mm_table_maker.py
#
# DESCRIPTION  : Upgraded version of mTable.py that allows to create several
#                multiplication tables at once.
#
# USAGE        : `$ python3/py mm_table.py 6 13 54 0 39`
#                (Creates 6x6, 13x13, 54x54, 0x0, 39x39 multiplication tables.)
#       
# AUTHOR       : Tim Kornev (@Timmate on GitHub)
#
# CREATED DATE : 9th of July, 2016
#


import sys

import openpyxl as xl
from openpyxl.styles import Font


def m_table(N):
    """Draws NxN multiplication table in an Excel spreasheet."""
    
    print()
    print('Creating a new workbook...')
    wb = xl.Workbook()
    sheet = wb.active
    
    # Draw label lines.
    print('Drawing label lines...')
    font_obj = Font(sz=13, bold=True)
    
    for row in range(2, N + 2):
        sheet['A' + str(row)].font = font_obj
        sheet['A' + str(row)] = row - 1

    for col in range(2, N + 2):
        sheet.cell(row=1, column=col).font = font_obj
        sheet.cell(row=1, column=col).value = col - 1

    # Draw the multiplication table.
    print('Drawing {}x{} multiplication table...'.format(N, N))
    for row in range(1, N + 1):
        for col in range(1, N + 1):
            sheet.cell(row=row+1, column=col+1).value = row * col
            
    wb.save('{}x{}_m_table.xlsx'.format(N, N))

# Take integers from command line arguments.
if len(sys.argv) >= 2:
    for argv in sys.argv[1:]:
        N = int(argv)
        m_table(N)
else:
    raise Exception('Invalid number of arguments.')

print()
print('Done.')
print()
