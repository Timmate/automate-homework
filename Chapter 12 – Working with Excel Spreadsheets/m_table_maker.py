#! python3
#
# NAME         : m_table_maker.py
#
# DESCRIPTION  : Creates a NxN multiplication table in an Excel spreadsheet.
#                The N is integer given by the user the command line
#                arguments.
#
# USAGE        : `$ python3/py m_table.py 6`
#
# AUTHOR       : Tim Kornev (@Timmate on GitHub)
#
# CREATED DATE : 9th of July, 2016
#


import sys

import openpyxl as xl
from openpyxl.styles import Font


print()
print('Creating a new workbook...')
wb = xl.Workbook()
sheet = wb.active

# Take integers from command line arguments.
if len(sys.argv) == 2:
    N = int(sys.argv[1])
else:
    raise Exception('Invalid number of arguments.')

# Draw label lines.
print('Drawing label lines...')
font_obj = Font(sz=13, bold=True)

for row in range(2, N + 2):
    sheet['A' + str(row)].font = font_obj
    sheet['A' + str(row)] = row - 1
    
for col in range(2, N + 2):
        sheet.cell(row=1, column=col).font = font_obj
        sheet.cell(row=1, column=col).value = col - 1

# Draw the multiplication table.
print('Drawing {}x{} multiplication table...'.format(N, N))

for row in range(1, N + 1):
    for col in range(1, N + 1):
        sheet.cell(row=row+1, column=col+1).value = row * col

wb.save('{}x{}_m_table.xlsx'.format(N, N))

print()
print('Done.')
print()

