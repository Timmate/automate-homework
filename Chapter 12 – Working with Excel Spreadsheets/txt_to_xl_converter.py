#! python3
#
# NAME         : txt_to_xl_converter.py
#
# DESCRIPTION  : Reads in the contents of several text files and inserts those
#                contents into an Excel spreadsheet with one text file per
#                column and one line of text per row.
#
# AUTHOR       : Tim Kornev (@Timmate on GitHub)
#
# CREATED DATE : 9th of July, 2016
#


import os

import openpyxl
from openpyxl.cell import get_column_letter


FILENAME = 'converted_from_txt.xlsx'    # spreadsheet's name


print()
print('SAVE AS: {}'.format(FILENAME))
print()
wb = openpyxl.Workbook()
sheet = wb.active

col_number = 1    # increment column's number for each text file

for filename in sorted(os.listdir('.')):
    # Read in text files.
    if filename.endswith('.txt'):
        print('Reading in {}...'.format(filename))
        with open(filename) as f:
            contents = f.readlines()
        
        number_of_lines = len(contents)    # count number of lines in text file
        lines_lengths = []    # store here lines' lengths
        
        # Write data to the Excel spreadsheet.
        for row in range(1, number_of_lines + 1):
            sheet.cell(row=row, column=col_number).value = contents[row - 1]
            
            lines_lengths.append(len(contents[row - 1]))

        # Find the longest line.
        max_length = max(lines_lengths)
        # Set column's width. (a bit of designing)
        col_letter = get_column_letter(col_number)
        sheet.column_dimensions[col_letter].width = max_length + 5
        
        col_number += 1

wb.save(FILENAME)

print()
print('Done.')
print()
