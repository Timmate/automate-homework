#! python3
#
# NAME         : cell_inverter.py
#
# DESCRIPTION  : Inverts row and column of cells in an Excel spreadsheet.
#
# USAGE        : `$ python3/py cell_inverter.py bob_eats_cats.xlsx`
#       
# AUTHOR       : Tim Kornev (@Timmate on GitHub)
#
# CREATED DATE : 9th of July, 2016
#


import sys

import openpyxl


if len(sys.argv) == 2:
    FILENAME = sys.argv[1]
    SAVE_AS = 'inverted_{}'.format(FILENAME)
else:
    raise Exception('Invalid number of arguments.')

print()
print('Reading in {}...'.format(FILENAME))
wb = openpyxl.load_workbook(FILENAME)
sheet = wb.active
new_wb = openpyxl.Workbook()
new_sheet = new_wb.active

# Invert the row and the column of the cells.
for row in range(1, sheet.max_row + 1):
    for col in range(1, sheet.max_column + 1):
        new_sheet.cell(row=col, column=row).value = sheet.cell(row=row, column=col).value

print('Saving new spreadsheet as {}...'.format(SAVE_AS))
new_wb.save(SAVE_AS)

print()
print('Done.')
print()
