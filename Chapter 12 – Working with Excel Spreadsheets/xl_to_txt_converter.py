#! python3
#
# NAME         : xl_to_txt_converter.py
#
# DESCRIPTION  : Reads in an Excel spreadsheet and writes its contents to text files,
#                so that each column of table is corresponding to its text file that
#                contains every row from a column.
#       
# AUTHOR       : Tim Kornev (@Timmate on GitHub)
#
# CREATED DATE : 9th of July, 2016
#


import os
import sys

import openpyxl
from openpyxl.cell import get_column_letter

if len(sys.argv) == 2:
    FILENAME = sys.argv[1]
else:
    raise Exception('Invalid number of arguments.')

print()
print('Reading in {}...'.format(FILENAME))
wb = openpyxl.load_workbook(FILENAME)
sheet = wb.active

filename_suffix = '_column.txt'

# Loop over each column.
for col in range(1, sheet.max_column + 1):
    col_letter = get_column_letter(col)
    print('Writing data to {}{}'.format(col_letter, filename_suffix))
    filename = open('{}{}'.format(col_letter, filename_suffix), 'w')

    # Loop over each row in the column.
    for row in range(1, sheet.max_row + 1):

        row_value = str(sheet[col_letter + str(row)].value)

        # Add newline character to row's value.
        if not row_value.endswith('\n'):
            row_value += '\n'

        # Write result to a file.
        if row_value is None:
            # Skip an empty row.
            continue
        else:
            filename.write(row_value)

    filename.close()

print()
print('Done.')
print()
