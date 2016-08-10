#! python3
#
# NAME         : blank_rows_inserter.py
#
# DESCRIPTION  : Inserts blank rows into an Excel spreasheet.
#                Note: It does not apply original file's formatting to a
#                new spreadsheet with inserted blank rows.
#
# USAGE        : `$ python3/py blank_rows_inserter.py 5 2 cat_tails_pricelist.xlsx`
#                (Inserts 2 blank rows starting from 5 row into 
#                 cat_tails_pricelist.xlsx's active spreadsheet.)
#
# AUTHOR       : Tim Kornev (@Timmate  on GitHub)
#
# CREATED DATE : 9th of July, 2016
#


import sys

import openpyxl


# Take command line arguments.
if len(sys.argv) == 4:
    START_ROW = int(sys.argv[1])
    BLANK_ROWS = int(sys.argv[2])
    FILENAME = sys.argv[3]
    SAVE_AS = 'blank_rows_{}'.format(FILENAME)
    
else:
    raise Exception('Invalid number of arguments.')

print()
print('Reading in {}...'.format(FILENAME))

# Read in data from the *FILENAME* spreadsheet and write new data with
# inserted blank rows to another spreadsheet.
wb = openpyxl.load_workbook(FILENAME)
sheet = wb.active
new_wb = openpyxl.Workbook()
new_sheet = new_wb.active

# Check whether *START_ROW* < existing max row.
assert START_ROW < sheet.max_row, 'Start row must come before an existing ' \
       'max row ({}).'.format(sheet.max_row)

print('Inserting blank rows...')

# Write data from those lines that come before *START_ROW* and
# *START_ROW* line itself.
for row in range(1, START_ROW + 1):
    for col in range(1, sheet.max_column + 1):
        new_sheet.cell(row=row, column=col).value = sheet.cell(row=row, column=col).value

# Write other data, skipping "blank rows".
for row in range(START_ROW + 1, sheet.max_row + 1):
    for col in range(1, sheet.max_column + 1):       
        new_row = row + BLANK_ROWS    # skip blank rows
        new_sheet.cell(row=new_row, column=col).value = sheet.cell(row=row, column=col).value

print('Saving new spreadsheet as {}...'.format(SAVE_AS))
new_wb.save(SAVE_AS)

print()
print('Done.')
print()
