#! python3
#
# NAME         : xl_to_csv_converter.py
#
# DESCRIPTION  : Converts each Excel workbook (including each spreadsheet) in
#                current working directory to CSV format so that each spreadsheet
#                is converted to a separate CSV file.
#
# AUTHOR       : Tim Kornev (@Timmate on GitHub)
#
# CREATED DATE : 9th of July, 2016
#


import os
import csv

import openpyxl
from openpyxl.cell import get_column_letter


SAVE_DIR = 'converted_excel_spreadsheets'
EXCEL_EXTENSIONS = ['.xlsx', '.xls', '.xml']

os.makedirs(SAVE_DIR, exist_ok=True)

print()
print('SAVE CONVERTED FILES IN: {}'.format(SAVE_DIR))

for filename in os.listdir('.'):
    # Read in Excel workbooks.
    for excel_extension in EXCEL_EXTENSIONS:    # check for each extension
        if filename.endswith(excel_extension)
            # Open the workbook.
            print('Converting {} workbook...'.format(filename))
            wb = openpyxl.load_workbook(filename)

            # Iterate over each sheet.
            for sheet in wb.get_sheet_names():
                print('Gathering data from {} sheet...'.format(sheet))
                active_sheet = wb.get_sheet_by_name(sheet)

                sheet_data = []   # contain sheet data here

                for row in range(1, active_sheet.max_row + 1):
                    
                    row_data = []    # contain row data here
                    
                    # Get data from a row.
                    for col in range(1, active_sheet.max_column + 1):
                
                        cell_value = active_sheet.cell(row=row, column=col)
                        row_data.append(cell_value)

                    # Add the row data to the sheet data.
                    sheet_data.append(row_data)

                # Write the sheet data to a CSV file.
                filename = filename.split('.')[0]
                save_path = os.path.join(SAVE_DIR, filename)
                write_file = open('{}_{}.csv'.format(save_path, sheet), 'w', newline='')
                csv_writer = csv.writer(write_file)
                csv_writer.writerows(sheet_data)
                write_file.close()

print()
print('Done.')
print()
